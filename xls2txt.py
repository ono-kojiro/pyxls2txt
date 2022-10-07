#!/usr/bin/env python3

import os
import sys

import getopt
import json

import re

import xlrd
import openpyxl
from openpyxl.utils import get_column_letter

from pprint import pprint

def usage():
	print("Usage : {0} [-o output.txt] [input.xls(x)]".format(sys.argv[0]))

def _xlsx2txt(fp, filepath)	:
	wb = openpyxl.load_workbook(filename = filepath)
	for sheet_name in wb.sheetnames:
		sheet = wb[sheet_name]
		fp.write("#   sheet : {0}\n".format(sheet_name))
		
		hidden_cols = {}
		for collet, coldim in sheet.column_dimensions.items():
			if coldim.hidden == True :
				hidden_cols[collet] = 1

		hidden_rows = {}
		for rowlet, rowdim in sheet.row_dimensions.items():
			if rowdim.hidden == True :
				hidden_rows[str(rowlet)] = 1;
		
		rows = len(tuple(sheet.rows))
		cols = len(tuple(sheet.columns))
		for row in range(rows) :
			if str(row) in hidden_rows :
				continue
			
			fp.write(" ")
			for col in range(cols):
				col_let = get_column_letter(col + 1)
				if col_let in hidden_cols:
					continue
				
				cell = sheet.cell(row=row + 1, column=col + 1)
				val = cell.value
				if col != 0 :
					fp.write("\t")
				if val == None:
					val = ''
				
				fp.write("{0}".format(val))
			fp.write("\n")

def _xls2txt(fp, filepath) :
	fp.write("# file : {0}\n".format(filepath))
		
	filename, ext = os.path.splitext(filepath)
		
	book = xlrd.open_workbook(filepath, formatting_info=True)

	for sheet in book.sheets() :
		fp.write("#   sheet : {0}\n".format(sheet.name))
			
		for row in range(sheet.nrows) :
			if row in sheet.rowinfo_map:
				if sheet.rowinfo_map[row].hidden == 1:
					continue
			
			fp.write(" ")
			for col in range(sheet.ncols) :
				if col in sheet.colinfo_map:
					if sheet.colinfo_map[col].hidden == 1:
						continue
					
				cell = sheet.cell(row, col)
				val  = cell.value
				if col != 0 :
					fp.write("\t")
				fp.write("{0}".format(val))
			fp.write("\n")

def xls2txt(fp, filepath) :
	fp.write("# file : {0}\n".format(filepath))
	filename, ext = os.path.splitext(filepath)
		
	if ext == ".xls" :
		_xls2txt(fp, filepath)
	elif ext == ".xlsx" :
		_xlsx2txt(fp, filepath)

def main():
	ret = 0

	try:
		opts, args = getopt.getopt(
			sys.argv[1:], "hvo:", ["help", "version", "output="])
	except getopt.GetoptError as err:
		print(str(err))
		sys.exit(2)
	
	output = None
	
	for o, a in opts:
		if o == "-v":
			usage()
			sys.exit(0)
		elif o in ("-h", "--help"):
			usage()
			sys.exit(0)
		elif o in ("-o", "--output"):
			output = a
		else:
			assert False, "unknown option"
	
	if ret != 0:
		sys.exit(1)

	if len(args) == 0 :
		usage()
		sys.exit(1)
	
	if output is not None :
		fp = open(output, mode='w', encoding='utf-8')
	else :
		fp = sys.stdout

	for filepath in args:
		#print("arg : {0}".format(filepath))
		fp.write("# file : {0}\n".format(filepath))
		
		filename, ext = os.path.splitext(filepath)
		
		if ext == ".xls" :
			book = xlrd.open_workbook(filepath, formatting_info=True)

			#pprint(book)
			for sheet in book.sheets() :
				#pprint(sheet)
				
				fp.write("#   sheet : {0}\n".format(sheet.name))
				#pprint(sheet.colinfo_map)
				#pprint(sheet.rowinfo_map)
				
				for row in range(sheet.nrows) :
					if row in sheet.rowinfo_map:
						if sheet.rowinfo_map[row].hidden == 1:
							continue
					
					fp.write(" ")
					for col in range(sheet.ncols) :
						if col in sheet.colinfo_map:
							if sheet.colinfo_map[col].hidden == 1:
								continue
							
						cell = sheet.cell(row, col)
						val  = cell.value
						if col != 0 :
							fp.write("\t")
						fp.write("{0}".format(val))
					fp.write("\n")
		elif ext == ".xlsx" :
			wb = openpyxl.load_workbook(filename = filepath)
			for sheet_name in wb.sheetnames:
				sheet = wb[sheet_name]
				fp.write("#   sheet : {0}\n".format(sheet_name))
				
				hidden_cols = {}
				for collet, coldim in sheet.column_dimensions.items():
					if coldim.hidden == True :
						hidden_cols[collet] = 1

				hidden_rows = {}
				for rowlet, rowdim in sheet.row_dimensions.items():
					if rowdim.hidden == True :
						hidden_rows[str(rowlet)] = 1;
				
				rows = len(tuple(sheet.rows))
				cols = len(tuple(sheet.columns))
				for row in range(rows) :
					if str(row) in hidden_rows :
						continue
					
					fp.write(" ")
					for col in range(cols):
						col_let = get_column_letter(col + 1)
						if col_let in hidden_cols:
							continue
						
						cell = sheet.cell(row=row + 1, column=col + 1)
						val = cell.value
						if col != 0 :
							fp.write("\t")
						if val == None:
							val = ''
						
						fp.write("{0}".format(val))
					fp.write("\n")
				#pprint(sheet)
			pass
	
	if output is not None :
		fp.close()
	
if __name__ == "__main__":
	main()
